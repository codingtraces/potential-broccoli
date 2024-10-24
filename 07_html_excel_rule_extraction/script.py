import os
from bs4 import BeautifulSoup
import openpyxl
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import chardet
import re

# Initialize logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

CATEGORIES = {
    "Queue": "Queue Rule",
    "Page": "Page Rule",
    "Component": "Component Rule",
    "Document": "Document Rule",
    "Function": "Function"
}


def detect_encoding(file_path):
    """Detect the encoding of the file."""
    try:
        with open(file_path, 'rb') as raw_file:
            raw_data = raw_file.read(1024 * 1024)  # Read first 1MB to detect encoding
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            logging.info(f"Detected encoding for {file_path}: {encoding}")
            return encoding
    except Exception as e:
        logging.error(f"Error detecting encoding: {e}")
        return 'utf-8'


def categorize_rule(rule_name):
    """Categorize the rule based on keywords."""
    for keyword, category in CATEGORIES.items():
        if keyword.lower() in rule_name.lower():
            return category
    return "Unknown"


def extract_rule_and_formula(soup):
    """Extract rules, formulas, and categorize them."""
    extracted_data = []

    # Find all sections containing rules
    rules = soup.find_all('div', class_='rule')

    for rule in rules:
        rule_header = rule.find('h3')
        formula = rule.find('div', class_='formula')

        if rule_header and formula:
            rule_name = rule_header.get_text(strip=True)
            category = categorize_rule(rule_name)
            formula_text = formula.get_text(separator="\n", strip=True)

            # Extract rule ID if present
            match = re.match(r"(R\d+|F\d+)\s*(.*)", rule_name)
            rule_id, rule_title = match.groups() if match else ("UNKNOWN", rule_name)

            # Store the extracted data
            extracted_data.append((rule_id, rule_title, formula_text, category))

    return extracted_data


def write_to_excel(extracted_data, output_excel):
    """Write the extracted data into an Excel file."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Extracted Rules and Formulas'

        # Write headers
        headers = ['Rule ID', 'Rule Name', 'Formula', 'Category']
        sheet.append(headers)

        for row_index, data in enumerate(extracted_data, start=2):
            sheet.append([data[0], data[1], data[2], data[3]])

            # Wrap text for the formula column
            formula_cell = sheet.cell(row=row_index, column=3)
            formula_cell.alignment = Alignment(wrap_text=True)

            # Auto-size columns
            for col in range(1, 5):
                sheet.column_dimensions[get_column_letter(col)].width = 50

        workbook.save(output_excel)
        logging.info(f"Excel file saved at {output_excel}")
    except Exception as e:
        logging.error(f"Error writing to Excel: {e}")


def process_html_file(input_file, output_excel):
    """Process a single HTML file and extract rules."""
    try:
        encoding = detect_encoding(input_file)
        logging.info(f"Processing {input_file}")

        with open(input_file, 'r', encoding=encoding, errors='ignore') as file:
            soup = BeautifulSoup(file, 'html.parser')
            extracted_data = extract_rule_and_formula(soup)

            if extracted_data:
                write_to_excel(extracted_data, output_excel)
            else:
                logging.warning("No data extracted from the provided HTML file.")
    except Exception as e:
        logging.error(f"Error processing {input_file}: {e}")


if __name__ == "__main__":
    input_file = './input_htm/sample.html'
    output_excel = './output/formula_report.xlsx'

    # Process the HTML and generate the Excel report
    try:
        process_html_file(input_file, output_excel)
    except Exception as e:
        logging.error(f"Critical error: {e}")
