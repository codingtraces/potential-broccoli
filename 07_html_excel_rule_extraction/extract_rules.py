import os
import re
import logging
import chardet
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def detect_encoding(file_path):
    """Detect encoding using chardet for robustness."""
    with open(file_path, 'rb') as file:
        raw_data = file.read(4096)
        result = chardet.detect(raw_data)
        encoding = result['encoding'] if result['confidence'] > 0.5 else 'utf-8'
        logging.info(f"Detected encoding: {encoding}")
    return encoding

def parse_rule_name(tag):
    """Extract rule ID and name from an h3 tag."""
    rule_pattern = re.match(r"^(R\d+)\s+(.*)", tag.get_text(strip=True))
    if rule_pattern:
        return rule_pattern.group(1), rule_pattern.group(2)
    return None, None

def categorize_rule(rule_name):
    """Categorize rules based on relevant keywords."""
    categories = {
        "Queue": "Queue Rule",
        "Page": "Page Rule",
        "Component": "Component",
        "Document": "Document Rule",
        "Design": "Page Design"
    }
    for keyword, category in categories.items():
        if keyword in rule_name:
            return category
    return "Uncategorized"

def extract_rules_from_html(file_path):
    """Extract rules from HTML files, handling encoding properly."""
    encoding = detect_encoding(file_path)
    with open(file_path, 'r', encoding=encoding, errors='replace') as file:
        soup = BeautifulSoup(file, 'html.parser')

    rules_data = []
    for rule in soup.find_all('div', class_='rule'):
        rule_heading = rule.find('h3')
        if rule_heading and not re.match(r"^F\d+", rule_heading.get_text(strip=True)):
            rule_id, rule_name = parse_rule_name(rule_heading)
            if rule_id:
                formula_text = rule.find('div', class_='formula').get_text('\n', strip=False)
                rules_data.append((
                    rule_id,
                    rule_name,
                    formula_text.strip(),
                    categorize_rule(rule_name)
                ))

    return rules_data

def auto_fit_columns(sheet):
    """Auto-fit Excel columns based on content width."""
    for col in sheet.columns:
        max_length = max(len(str(cell.value) or "") for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

def write_to_excel(data, output_path):
    """Write rules into an Excel sheet with preserved formatting."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Rules'

    # Add headers
    headers = ['Rule ID', 'Rule Name', 'Formula', 'Category']
    sheet.append(headers)

    # Populate rows
    for row_index, row in enumerate(data, start=2):
        sheet.append(row)

        # Apply text wrapping and alignment for the formula column
        formula_cell = sheet.cell(row=row_index, column=3)
        formula_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        formula_cell.font = Font(name='Courier New')  # For code-like formatting

    auto_fit_columns(sheet)
    workbook.save(output_path)
    logging.info(f"Rules report saved at {output_path}")

def process_rules(input_folder, output_file):
    """Extract and process all rules from HTML files."""
    all_rules = []
    total_files = sum(len(files) for _, _, files in os.walk(input_folder))
    processed_files = 0

    for root, _, files in os.walk(input_folder):
        for file in files:
            if file.endswith(('.html', '.htm')):
                file_path = os.path.join(root, file)
                logging.info(f"Processing: {file_path}")
                extracted_rules = extract_rules_from_html(file_path)
                all_rules.extend(extracted_rules)
                processed_files += 1
                logging.info(f"Progress: {processed_files}/{total_files} files processed")

    if all_rules:
        write_to_excel(all_rules, output_file)
    else:
        logging.warning("No rules found in the provided HTML files.")

if __name__ == "__main__":
    # Paths for input HTML folder and output Excel file
    input_folder = './input_htm'
    output_file = './output/rules_report.xlsx'

    process_rules(input_folder, output_file)
