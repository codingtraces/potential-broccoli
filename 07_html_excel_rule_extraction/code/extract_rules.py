import re
import os
import pandas as pd
from bs4 import BeautifulSoup
import html2text
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Category mapping to classify rules by keywords
CATEGORY_MAPPING = {
    'queue': 'Queue Rule',
    'component': 'Component Rule',
    'page': 'Page Layout / Design Rule',
    'banner': 'Banner Configuration',
    'document': 'Document Management',
}

def parse_html_in_chunks(file_path):
    """Reads large HTML files in manageable chunks."""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            soup = BeautifulSoup(file, 'lxml')

        # Extract only relevant tags to reduce noise
        content_blocks = soup.find_all(['pre', 'div', 'em'])
        for block in content_blocks:
            yield html2text.html2text(str(block))
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        return []

def extract_rules_and_formulas(content):
    """Extracts all rules and formulas between rule identifiers."""
    rule_pattern = r'(R\d+|F\d+)\s+([\w_]+)\s*(.*?)\s*(?=R\d+|F\d+|\Z)'
    matches = re.findall(rule_pattern, content, re.DOTALL)

    extracted_data = []
    for rule_id, rule_name, formula_content in matches:
        formula = extract_formula(formula_content)
        category = categorize_rule(rule_name)
        extracted_data.append({
            'Rule ID': rule_id,
            'Rule Name': rule_name,
            'Formula': formula,
            'Category': category
        })
    return extracted_data

def extract_formula(content):
    """Extracts the formula section from a block."""
    match = re.search(r'Formula:\s*(.*?)\s*$', content, re.DOTALL)
    return match.group(1).strip() if match else 'N/A'

def categorize_rule(rule_name):
    """Categorizes rules based on keywords in their names."""
    rule_name = rule_name.lower()
    for keyword, category in CATEGORY_MAPPING.items():
        if keyword in rule_name:
            return category
    return 'General Business Rule'

def extract_rules_from_html_file(input_file):
    """Processes a single HTML file and extracts all relevant data."""
    extracted_data = []
    try:
        for chunk in parse_html_in_chunks(input_file):
            extracted_data.extend(extract_rules_and_formulas(chunk))
    except Exception as e:
        print(f"Error extracting rules from {input_file}: {e}")
    return extracted_data

def extract_rules_from_folder(input_folder, output_file):
    """Processes all HTML files in the input folder."""
    all_data = []
    for filename in os.listdir(input_folder):
        if filename.endswith('.html'):
            input_file = os.path.join(input_folder, filename)
            print(f"Processing {input_file}...")
            all_data.extend(extract_rules_from_html_file(input_file))

    # Create DataFrame and save to Excel
    df = pd.DataFrame(all_data, columns=['Rule ID', 'Rule Name', 'Formula', 'Category'])
    df.to_excel(output_file, index=False)
    print(f"Rules extracted and saved to {output_file}")
    apply_wrap_text(output_file)

def apply_wrap_text(file_path):
    """Applies wrap text to the Formula column in the Excel file."""
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(file_path)
    print("Wrap text applied to the Formula column.")

# Define folder paths
input_folder = '../input'
output_folder = '../output'
output_file = os.path.join(output_folder, 'extracted_rules.xlsx')

if __name__ == '__main__':
    extract_rules_from_folder(input_folder, output_file)
