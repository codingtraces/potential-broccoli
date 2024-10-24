import os
from bs4 import BeautifulSoup
import openpyxl
import logging
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import chardet

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


def detect_encoding(file_path):
    """Detect encoding using chardet."""
    with open(file_path, 'rb') as f:
        raw_data = f.read(1024 * 1024)  # Read 1MB to detect encoding
        result = chardet.detect(raw_data)
        encoding = result.get('encoding', 'utf-8')
        logging.info(f"Detected encoding for {file_path}: {encoding}")
        return encoding


def extract_formula_content(tag):
    """Preserve indentation and formatting of formulas."""
    return tag.get_text(separator='\n', strip=False)


def extract_functions(soup):
    """Extract all functions from the HTML soup."""
    functions = []
    function_tags = soup.find_all('div', class_='rule')

    for tag in function_tags:
        header = tag.find('h3')
        if header and header.get_text(strip=True).startswith('F'):
            func_name = header.get_text(strip=True)
            formula_tag = tag.find('div', class_='formula')
            if formula_tag:
                formula = extract_formula_content(formula_tag)
                functions.append((func_name, formula))

    return functions


def adjust_column_width(sheet):
    """Adjust Excel column width based on content."""
    for col in sheet.columns:
        max_length = max(len(str(cell.value) or '') for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)


def write_functions_to_excel(functions, output_path):
    """Write functions data to Excel with proper formatting."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Functions'

    # Write headers
    sheet.append(['Function Name', 'Formula'])

    # Write function data
    for func in functions:
        sheet.append(func)
        formula_cell = sheet.cell(row=sheet.max_row, column=2)
        formula_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        formula_cell.font = Font(name='Courier New')  # Code-like font

    adjust_column_width(sheet)
    workbook.save(output_path)
    logging.info(f"Functions report saved at {output_path}")


def process_html_files(input_dir, output_path):
    """Process all HTML files to extract functions."""
    all_functions = []

    files = [f for f in os.listdir(input_dir) if f.endswith(('.html', '.htm'))]
    total_files = len(files)

    for i, file in enumerate(files, 1):
        file_path = os.path.join(input_dir, file)
        encoding = detect_encoding(file_path)

        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                soup = BeautifulSoup(f, 'html.parser')
                functions = extract_functions(soup)
                all_functions.extend(functions)

                progress = (i / total_files) * 100
                logging.info(f"Processed {i}/{total_files} files ({progress:.2f}% complete)")
        except Exception as e:
            logging.error(f"Error processing {file_path}: {e}")

    if all_functions:
        write_functions_to_excel(all_functions, output_path)
    else:
        logging.warning("No functions found in the provided HTML files.")


if __name__ == '__main__':
    # Define input and output paths
    input_directory = './input_htm'
    output_excel = './output/functions_report.xlsx'

    process_html_files(input_directory, output_excel)
