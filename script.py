import os
from bs4 import BeautifulSoup
import openpyxl
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import chardet

# Initialize logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def detect_encoding(file_path):
    """Detect the encoding of the file."""
    with open(file_path, 'rb') as raw_file:
        raw_data = raw_file.read()
        result = chardet.detect(raw_data)
        return result['encoding']

def extract_data_from_htm(file_path):
    extracted_data = []
    try:
        # Detect the file's encoding
        encoding = detect_encoding(file_path)
        logging.info(f"Detected encoding for {file_path}: {encoding}")

        # Try opening the file with the detected encoding
        try:
            with open(file_path, 'r', encoding=encoding) as file:
                soup = BeautifulSoup(file, 'html.parser')
                # Log the first 1000 characters to check if reading works
                logging.info(f"First 1000 characters of {file_path}: {soup.prettify()[:1000]}")
        except Exception as e:
            logging.warning(f"Failed to parse {file_path} with detected encoding {encoding}. Falling back to utf-8.")
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                soup = BeautifulSoup(file, 'html.parser')
                # Log the first 1000 characters to check if reading works with fallback
                logging.info(f"First 1000 characters of {file_path} (utf-8 fallback): {soup.prettify()[:1000]}")

        # Locate all the <pre> tags that contain the formulas
        formula_sections = soup.find_all('pre')

        for formula_tag in formula_sections:
            # Check if the formula is prefixed by "Formula:" to ensure relevance
            if 'Formula:' in formula_tag.text:
                # Find the nearest preceding <a> tag for the rule identifier
                previous_a_tag = formula_tag.find_previous('a')
                if previous_a_tag:
                    rule_title = previous_a_tag.text.strip()  # The rule title from the <a> tag
                    formula_text = formula_tag.get_text().strip()  # Extract the formula text
                    extracted_data.append((rule_title, formula_text, file_path))
                else:
                    logging.warning(f"No <a> tag found for formula in {file_path}")
    
    except Exception as e:
        logging.error(f"Error parsing file {file_path}: {e}")
    return extracted_data

def write_to_excel(extracted_data, output_excel):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Extracted Formulas'

        # Write headers
        headers = ['Rule Description', 'Formula', 'Source File']
        sheet.append(headers)

        for row_index, data in enumerate(extracted_data, start=2):  # start=2 to account for headers in row 1
            sheet.append(data)
            
            # Set formula cell to wrap text and preserve line breaks
            formula_cell = sheet.cell(row=row_index, column=2)  # Column 2 is for formula
            formula_cell.alignment = Alignment(wrap_text=True)

            # Optionally, auto-size the columns
            for col in range(1, 4):
                sheet.column_dimensions[get_column_letter(col)].width = 50

        # Save the Excel file
        workbook.save(output_excel)
        logging.info(f"Excel file saved at {output_excel}")
    except Exception as e:
        logging.error(f"Error writing to Excel: {e}")

def process_htm_reports(input_directory, output_excel):
    all_extracted_data = []
    for root, dirs, files in os.walk(input_directory):
        for file in files:
            if file.endswith(".htm"):
                file_path = os.path.join(root, file)
                extracted_data = extract_data_from_htm(file_path)
                all_extracted_data.extend(extracted_data)

    if all_extracted_data:
        write_to_excel(all_extracted_data, output_excel)
    else:
        logging.warning("No data extracted from the provided .htm files.")

# Example usage
if __name__ == "__main__":
    input_directory = './input_htm'
    output_excel = './output/formula_report.xlsx'

    process_htm_reports(input_directory, output_excel)
