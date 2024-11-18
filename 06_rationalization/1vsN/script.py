import os
import fitz  # PyMuPDF for analyzing structure and layout
from datetime import datetime
from collections import defaultdict
import pandas as pd
import concurrent.futures
import traceback
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
import re

# Helper function to normalize text for comparison
def normalize_text(text):
    text = text.lower().strip()
    text = re.sub(r'\b[xX]{5,}\b', '', text)  # Remove long patterns like xxxxxxx
    return text

# Function to validate PDF files
def validate_pdf(file_path):
    try:
        with fitz.open(file_path) as doc:
            if doc.is_encrypted:
                print(f"PDF {file_path} is encrypted. Skipping...")
                return False
            return True
    except Exception as e:
        print(f"Error opening {file_path}: {e}. Skipping...")
        return False

# Function to analyze a single PDF structure (text)
def analyze_pdf(file_path):
    try:
        with fitz.open(file_path) as doc:
            if doc.is_encrypted:
                print(f"PDF {file_path} is encrypted. Skipping...")
                return None

            text_blocks = []
            for page_number in range(len(doc)):
                page = doc.load_page(page_number)
                blocks = page.get_text("dict")["blocks"]
                for block in blocks:
                    if "lines" in block:
                        block_text = "\n".join([span["text"] for line in block["lines"] for span in line["spans"]]).strip()
                        block_text = normalize_text(block_text)

                        # Filter out small blocks of text less than 10 words
                        if len(block_text.split()) >= 10:
                            text_blocks.append(block_text)

            return {"text_blocks": text_blocks}
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        traceback.print_exc()
        return None

# Function to compare PDFs and find common elements with similarity percentages
def compare_pdf_structures(pdf_reports, single_pdf_report):
    common_elements = {"text_blocks": defaultdict(list)}
    all_text_blocks = single_pdf_report["text_blocks"] + [block for report in pdf_reports.values() for block in report["text_blocks"]]
    pdf_names = ["Single PDF"] * len(single_pdf_report["text_blocks"]) + [pdf_name for pdf_name, report in pdf_reports.items() for _ in report["text_blocks"]]

    vectorizer = TfidfVectorizer().fit_transform(all_text_blocks)
    similarity_matrix = cosine_similarity(vectorizer)

    for i in range(len(single_pdf_report["text_blocks"])):
        for j in range(len(single_pdf_report["text_blocks"]), len(all_text_blocks)):
            similarity = similarity_matrix[i, j]
            if similarity > 0.1:
                common_elements["text_blocks"][single_pdf_report["text_blocks"][i]].append((pdf_names[j], round(similarity * 100, 2)))

    return common_elements

# Function to generate HTML report for common elements in tabular format
def generate_comparison_html_report(common_elements, output_folder):
    html_content = f"""
    <html>
    <head>
        <title>Template Reusability Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; word-wrap: break-word; white-space: pre-wrap; }}
            th {{ background-color: #f2f2f2; }}
            .highlight-green {{ background-color: #d4edda; }}
            .highlight-yellow {{ background-color: #fff3cd; }}
        </style>
    </head>
    <body>
        <h1>Template Reusability Report</h1>
        <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <table>
            <thead>
                <tr>
                    <th>Type</th>
                    <th>Content (Paragraph)</th>
                    <th>Found in PDF</th>
                    <th>Similarity Percentage</th>
                </tr>
            </thead>
            <tbody>
    """

    for item, matches in common_elements["text_blocks"].items():
        for match in matches:
            pdf_name, similarity = match
            highlight_class = "highlight-green" if similarity == 100 else "highlight-yellow"
            html_content += f"""
            <tr class="{highlight_class}">
                <td>Text Block</td>
                <td>{item}</td>
                <td>{pdf_name}</td>
                <td>{similarity}%</td>
            </tr>
            """

    html_content += """
            </tbody>
        </table>
    </body>
    </html>
    """

    html_filename = os.path.join(output_folder, "template_reusability_report.html")
    with open(html_filename, "w", encoding="utf-8") as html_file:
        html_file.write(html_content)

    print(f"Template reusability report generated: {html_filename}")

# Function to generate Excel report with wrapped text and highlighting
def generate_comparison_excel_report(common_elements, output_folder):
    rows = []
    for item, matches in common_elements["text_blocks"].items():
        for match in matches:
            pdf_name, similarity = match
            rows.append(["Text Block", item, pdf_name, similarity])

    df = pd.DataFrame(rows, columns=["Type", "Content (Paragraph)", "Found in PDF", "Similarity Percentage"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Reusability Report"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if r_idx > 1:
                if c_idx == 4:
                    similarity = float(value)
                    fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") if similarity == 100 else PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    ws.cell(row=r_idx, column=2).fill = fill
                    ws.cell(row=r_idx, column=4).fill = fill

    current_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
    excel_filename = os.path.join(output_folder, f"template_reusability_report_{current_time}.xlsx")
    wb.save(excel_filename)

    print(f"Excel report generated: {excel_filename}")

# Main function to process the PDF analysis and comparison
def analyze_single_vs_all(single_pdf_folder, all_pdf_folder, base_output_folder):
    single_pdf_files = [f for f in os.listdir(single_pdf_folder) if f.lower().endswith('.pdf')]
    if not single_pdf_files:
        print("No valid PDF files found in the singlepdf folder.")
        return

    single_pdf = os.path.join(single_pdf_folder, single_pdf_files[0])
    if not validate_pdf(single_pdf):
        print(f"Single PDF {single_pdf} is not valid. Skipping analysis.")
        return

    all_pdf_files = [os.path.join(all_pdf_folder, f) for f in os.listdir(all_pdf_folder) if f.lower().endswith('.pdf')]
    all_pdf_files = [pdf for pdf in all_pdf_files if validate_pdf(pdf)]

    if not all_pdf_files:
        print("No valid PDF files found in the allpdf folder.")
        return

    print(f"Analyzing single PDF: {single_pdf}")
    single_pdf_report = analyze_pdf(single_pdf)
    if single_pdf_report is None:
        print("Failed to process the single PDF.")
        return

    pdf_reports = {}
    processed_count = 0
    total_pdfs = len(all_pdf_files)
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {executor.submit(analyze_pdf, pdf): pdf for pdf in all_pdf_files}
        for future in concurrent.futures.as_completed(futures):
            pdf = futures[future]
            try:
                report = future.result()
                if report:
                    pdf_reports[os.path.basename(pdf)] = report
                processed_count += 1
                print(f"Processed {processed_count}/{total_pdfs} PDFs...")
            except Exception as exc:
                print(f"PDF {pdf} generated an exception: {exc}")
                with open(os.path.join(base_output_folder, "processing_log.txt"), 'a') as log_file:
                    log_file.write(f"{pdf} failed with error: {exc}\n")

    current_time = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
    output_folder = os.path.join(base_output_folder, f"pdf_rationalization_report_{current_time}")
    os.makedirs(output_folder, exist_ok=True)

    common_elements = compare_pdf_structures(pdf_reports, single_pdf_report)
    generate_comparison_html_report(common_elements, output_folder)
    generate_comparison_excel_report(common_elements, output_folder)

if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    single_pdf_folder = os.path.join(base_dir, 'singlepdf')
    all_pdf_folder = os.path.join(base_dir, 'allpdf')
    base_output_folder = os.path.join(base_dir, 'result')

    os.makedirs(base_output_folder, exist_ok=True)

    print("Starting analysis...")
    analyze_single_vs_all(single_pdf_folder, all_pdf_folder, base_output_folder)
    print("Analysis complete.")
